"""Convert VT DFR quarterly xlsx → CSV (utf-8). Usage: python xlsx-to-csv.py in.xlsx out.csv"""
import csv
import sys
from openpyxl import load_workbook

def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: python xlsx-to-csv.py <in.xlsx> <out.csv>", file=sys.stderr)
        return 1
    src, dest = sys.argv[1], sys.argv[2]
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        print("empty workbook", file=sys.stderr)
        return 1
    header_names = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers)]
    with open(dest, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header_names)
        for row in rows:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            out = []
            for c in row:
                if c is None:
                    out.append("")
                elif hasattr(c, "isoformat"):
                    out.append(c.isoformat()[:10])
                else:
                    out.append(str(c).replace("\r", " ").replace("\n", " ").strip())
            w.writerow(out)
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
